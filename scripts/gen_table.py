"""G1レースの出走馬分析表を生成するスクリプト。

コマンド:
cd path/to/g1-predict
python -m scripts.gen_table --race-code <16桁 race_code>
"""

import argparse
import os
from typing import Any

import yaml
from dotenv import find_dotenv, load_dotenv
from keiba_data_interface import DataInterface
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from g1_predict.modules.gen_table.table_context import TableContext
from g1_predict.modules.gen_table.table_utils import (
    HEADER_FILL,
    WAKU_FILLS,
    apply_color_rules,
    is_na,
    to_cell_value,
)

load_dotenv(find_dotenv())

_REPO_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
_CONFIGS_DIR = os.path.join(_REPO_DIR, "configs")
_PUBLIC_DIR = os.path.join(_REPO_DIR, "public")

_FIXED_COLS = [
    {"name": "枠", "source": {"type": "entry_field", "field": "枠番"}},
    {"name": "馬番", "source": {"type": "entry_field", "field": "馬番"}},
    {"name": "馬名", "source": {"type": "entry_field", "field": "馬名"}},
]


def generate_table(race_code: str) -> None:
    """指定レースの分析表を生成する。

    Args:
        race_code (str): 16桁レースコード。
    """
    di = DataInterface("mykeibadb")
    race_info = di.get_race_basic_info(race_code)
    race_name = str(race_info["競走名本題"].iloc[0])
    race_year = int(str(race_info["開催年"].iloc[0]))

    config_path = os.path.join(_CONFIGS_DIR, f"{race_name}.yml")
    with open(config_path, encoding="utf-8") as f:
        config = yaml.safe_load(f)

    entry_df = di.get_entry(race_code)
    ctx = TableContext(race_code, race_year, race_name, entry_df, di)

    sheet_data = []
    for sheet_name, columns in config["table"].items():
        col_configs = _FIXED_COLS + columns

        rows = []
        for _, horse in entry_df.iterrows():
            horse_id = str(horse["血統登録番号"]).strip()
            row: dict[str, Any] = {}
            for col_cfg in col_configs:
                row[col_cfg["name"]] = ctx.get_value(horse, horse_id, col_cfg["source"])
            rows.append(row)

        sheet_data.append((sheet_name, col_configs, rows))

    _write_xlsx(sheet_data, race_code, race_name, race_year)


def main() -> None:
    """エントリポイント。"""
    parser = argparse.ArgumentParser(description="G1レース分析表を生成する")
    parser.add_argument("--race-code", required=True, help="16桁 race_code")
    args = parser.parse_args()
    generate_table(args.race_code)


def _write_xlsx(
    sheet_data: list[tuple[str, list[dict[str, Any]], list[dict[str, Any]]]],
    race_code: str,
    race_name: str,
    race_year: int,
) -> str:
    """分析表データをxlsxファイルに書き出す。

    Args:
        sheet_data (list[tuple[str, list[dict[str, Any]], list[dict[str, Any]]]]):
            (シート名, カラム設定リスト, 行データリスト) のタプルのリスト。
        race_code (str): 16桁レースコード。出力ファイル名に使用。
        race_name (str): レース名。出力先ディレクトリ・ファイル名に使用。
        race_year (int): レース開催年。出力先ディレクトリに使用。

    Returns:
        str: 生成したxlsxファイルの絶対パス。
    """
    table_dir = os.path.join(_PUBLIC_DIR, str(race_year), f"{race_code}_{race_name}", "table")
    os.makedirs(table_dir, exist_ok=True)
    output_path = os.path.join(table_dir, f"{race_code}_{race_name}.xlsx")

    wb = Workbook()
    wb.remove(wb.active)

    for sheet_name, col_configs, rows in sheet_data:
        ws = wb.create_sheet(title=sheet_name)
        _write_sheet(ws, col_configs, rows)

    wb.save(output_path)
    print(f"生成完了: {output_path}")
    return output_path


def _write_sheet(
    ws: Worksheet,
    col_configs: list[dict[str, Any]],
    rows: list[dict[str, Any]],
) -> None:
    """1シート分の分析表データを書き出す。

    Args:
        ws (Worksheet): 書き込み対象のワークシート。
        col_configs (list[dict[str, Any]]): カラム設定リスト
            （name/source/color_rules等を持つ辞書）。
        rows (list[dict[str, Any]]): 馬ごとのカラム名→値の辞書リスト。
    """
    headers = [c["name"] for c in col_configs]

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = Font(bold=True)

    for row_idx, row in enumerate(rows, 2):
        for col_idx, col_cfg in enumerate(col_configs, 1):
            col_name = col_cfg["name"]
            value = row.get(col_name)

            display_map = col_cfg.get("display_map")
            display_value = value
            if display_map is not None and value is not None and not is_na(value):
                display_value = display_map.get(str(value), value)

            cell = ws.cell(row=row_idx, column=col_idx, value=to_cell_value(display_value))

            if col_name == "枠" and value is not None and not is_na(value):
                waku_fill = WAKU_FILLS.get(int(value))
                if waku_fill:
                    cell.fill = waku_fill
            else:
                color_rules = col_cfg.get("color_rules", [])
                if color_rules:
                    fill = apply_color_rules(value, color_rules)
                    if fill:
                        cell.fill = fill

    ws.freeze_panes = "A2"

    for col_idx, col_cfg in enumerate(col_configs, 1):
        col_letter = get_column_letter(col_idx)
        max_len = len(str(col_cfg["name"]))
        for row_idx in range(2, len(rows) + 2):
            cell_val = ws.cell(row=row_idx, column=col_idx).value
            if cell_val is not None:
                max_len = max(max_len, len(str(cell_val)))
        ws.column_dimensions[col_letter].width = max_len + 2


if __name__ == "__main__":
    main()

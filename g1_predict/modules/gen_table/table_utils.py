"""テーブル生成で使うユーティリティ関数と定数。"""

from datetime import date
from typing import Any

import pandas as pd
from openpyxl.styles import PatternFill

_COLOR_FILLS: dict[str, PatternFill] = {
    "green": PatternFill(patternType="solid", fgColor="FF92D050"),
    "yellow": PatternFill(patternType="solid", fgColor="FFFFEB9C"),
    "blue": PatternFill(patternType="solid", fgColor="FF9DC3E6"),
    "red": PatternFill(patternType="solid", fgColor="FFFF9999"),
    "orange": PatternFill(patternType="solid", fgColor="FFFFC000"),
}

HEADER_FILL = PatternFill(patternType="solid", fgColor="FFD9D9D9")

WAKU_FILLS: dict[int, PatternFill] = {
    1: PatternFill(patternType="solid", fgColor="FFFFFFFF"),
    2: PatternFill(patternType="solid", fgColor="FF444444"),
    3: PatternFill(patternType="solid", fgColor="FFE95556"),
    4: PatternFill(patternType="solid", fgColor="FF416BBA"),
    5: PatternFill(patternType="solid", fgColor="FFE7C52C"),
    6: PatternFill(patternType="solid", fgColor="FF45AF4C"),
    7: PatternFill(patternType="solid", fgColor="FFEE9738"),
    8: PatternFill(patternType="solid", fgColor="FFEF8FA0"),
}

SHIBA_TRACK_CODES: frozenset[str] = frozenset(
    {str(c) for c in range(10, 23)} | {str(c) for c in range(51, 60)}
)
DIRT_TRACK_CODES: frozenset[str] = frozenset({str(c) for c in range(23, 30)})


def is_na(value: Any) -> bool:
    """値がNaN/Noneに相当するか判定する。

    Args:
        value (Any): 判定する値。

    Returns:
        bool: NaN/Noneの場合True、それ以外はFalse。
    """
    if value is None:
        return True
    try:
        result = pd.isna(value)
        return bool(result)
    except (TypeError, ValueError):
        return False


def apply_color_rules(value: Any, color_rules: list[dict[str, Any]]) -> PatternFill | None:
    """カラールール設定に従い最初に一致したPatternFillを返す。

    Args:
        value (Any): 評価する値。
        color_rules (list[dict[str, Any]]): YAMLのcolor_rules設定リスト。

    Returns:
        PatternFill | None: 最初に条件が一致したルールのPatternFill。一致しない場合はNone。
    """
    for rule in color_rules:
        cond = rule["condition"]
        if _match_op(value, cond["op"], cond["value"]):
            return _COLOR_FILLS.get(rule["color"])
    return None


def filter_df(df: pd.DataFrame, filters: list[dict[str, Any]]) -> pd.DataFrame:
    """フィルタ設定リストをDataFrameに順次適用して絞り込む。

    Args:
        df (pd.DataFrame): 絞り込む対象のDataFrame。
        filters (list[dict[str, Any]]): YAMLのfilters設定リスト（field/op/valueキーを持つ辞書のリスト）。

    Returns:
        pd.DataFrame: フィルタ適用後のDataFrame。
    """
    for filt in filters:
        field, op, threshold = filt["field"], filt["op"], filt["value"]
        if field not in df.columns:
            return pd.DataFrame(columns=df.columns)
        mask = df[field].apply(lambda v, o=op, t=threshold: _match_op(v, o, t))
        df = df[mask]
    return df


def filter_by_horse(df: pd.DataFrame, horse_id: str) -> pd.DataFrame:
    """DataFrameをketto_toroku_bangoで絞り込む。

    Args:
        df (pd.DataFrame): 絞り込む対象のDataFrame。
        horse_id (str): 血統登録番号。

    Returns:
        pd.DataFrame: 絞り込み後のDataFrame。
    """
    return df[df["ketto_toroku_bango"].astype(str).str.strip() == horse_id]


def stat_result(stat: str, wins: int, top3: int, total: int) -> Any:
    """stat種別に対応する統計値を計算して返す。

    Args:
        stat (str): 統計種別。"wins" / "top3" / "win_rate" / "top3_rate" のいずれか。
        wins (int): 1着回数。
        top3 (int): 3着以内回数。
        total (int): 総出走数。

    Returns:
        Any: int（"wins"/"top3"）またはfloat（"win_rate"/"top3_rate"）。不明なstatの場合はNone。
    """
    if stat == "wins":
        return wins
    if stat == "top3":
        return top3
    if stat == "win_rate":
        return round(wins / total, 4) if total > 0 else 0.0
    if stat == "top3_rate":
        return round(top3 / total, 4) if total > 0 else 0.0
    return None


def aggregate_stat(df: pd.DataFrame, stat: str) -> Any:
    """DataFrameの確定着順からwins/top3/totalを集計して統計値を返す。

    Args:
        df (pd.DataFrame): kakutei_chakujunカラムを持つ馬ごとデータ。
        stat (str): 統計種別。stat_result参照。

    Returns:
        Any: intまたはfloatの統計値。stat_result参照。
    """
    kakutei = pd.to_numeric(
        df.get("kakutei_chakujun", pd.Series(dtype="float64")), errors="coerce"
    )
    total = len(df)
    wins = int((kakutei == 1).sum())
    top3 = int(((kakutei >= 1) & (kakutei <= 3)).sum())
    return stat_result(stat, wins, top3, total)


def to_cell_value(v: Any) -> Any:
    """NaN/NoneはNoneに、それ以外はvをそのまま返す。

    Args:
        v (Any): セルに書き込む値。

    Returns:
        Any: NaN/NoneはNone、それ以外はv。
    """
    if is_na(v):
        return None
    return v


def year_range(race_year: int, years: int) -> tuple[date, date]:
    """過去years年分の日付範囲を返す。

    Args:
        race_year (int): 基準年（レース開催年）。
        years (int): 遡る年数。

    Returns:
        date: 開始日（race_year - years + 1年の1月1日）。
        date: 終了日（race_year年の12月31日）。
    """
    return date(race_year - years + 1, 1, 1), date(race_year, 12, 31)


def _match_op(value: Any, op: str, threshold: Any) -> bool:
    """演算子を使ってvalueとthresholdを比較する。

    Args:
        value (Any): 比較対象の値。
        op (str): 演算子文字列（"==" / "!=" / ">=" / "<=" / ">" / "<" / "in" / "not_in" / "contains"）。
        threshold (Any): 比較する基準値。

    Returns:
        bool: 比較結果。valueがNaN、またはTypeError/ValueErrorが発生した場合はFalse。

    Raises:
        ValueError: opが不明な演算子文字列の場合。
    """
    if is_na(value):
        return False
    try:
        if op == "==":
            return bool(value == threshold)
        if op == "!=":
            return bool(value != threshold)
        if op == ">=":
            return bool(value >= threshold)
        if op == "<=":
            return bool(value <= threshold)
        if op == ">":
            return bool(value > threshold)
        if op == "<":
            return bool(value < threshold)
        if op == "in":
            return value in threshold
        if op == "not_in":
            return value not in threshold
        if op == "contains":
            return str(threshold) in str(value)
    except (TypeError, ValueError):
        return False
    raise ValueError(f"不明なop: {op}")

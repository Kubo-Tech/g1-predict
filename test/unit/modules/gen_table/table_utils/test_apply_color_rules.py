"""apply_color_rules の単体テスト。"""
from typing import Any

import pytest
from openpyxl.styles import PatternFill

from g1_predict.modules.gen_table.table_utils import apply_color_rules


def _rule(op: str, value: Any, color: str) -> dict[str, Any]:
    return {"condition": {"op": op, "value": value}, "color": color}


# 正常系
def test_apply_color_rules_returns_fill_when_matches() -> None:
    """条件に一致したルールのPatternFillを返す。"""
    rules = [_rule("==", 1, "green")]
    result = apply_color_rules(1, rules)
    assert isinstance(result, PatternFill)


def test_apply_color_rules_returns_none_when_no_match() -> None:
    """一致するルールがないときNoneを返す。"""
    rules = [_rule("==", 99, "green")]
    assert apply_color_rules(1, rules) is None


def test_apply_color_rules_returns_none_when_rules_empty() -> None:
    """ルールが空のときNoneを返す。"""
    assert apply_color_rules(1, []) is None


def test_apply_color_rules_returns_first_matching_rule() -> None:
    """複数ルールが一致する場合、最初のルールを返す。"""
    rules = [_rule(">=", 1, "green"), _rule(">=", 1, "red")]
    result = apply_color_rules(1, rules)
    assert result is not None
    assert result.fgColor.rgb == "FF92D050"  # green


def test_apply_color_rules_returns_none_for_na_value() -> None:
    """NA値はどのルールにも一致せずNoneを返す。"""
    rules = [_rule("==", None, "green")]
    assert apply_color_rules(None, rules) is None


@pytest.mark.parametrize(
    "value, op, threshold, color",
    [
        (0.3, ">=", 0.3, "green"),
        (0.2, "<", 0.3, "red"),
        ("A", "==", "A", "blue"),
        (5, "in", [3, 5, 7], "yellow"),
    ],
)
def test_apply_color_rules_various_conditions(
    value: Any, op: str, threshold: Any, color: str
) -> None:
    """様々な演算子とデータ型でPatternFillが返る。"""
    rules = [_rule(op, threshold, color)]
    result = apply_color_rules(value, rules)
    assert isinstance(result, PatternFill)


@pytest.mark.parametrize(
    "value",
    ["G1 9着", "G2 2着", "G3 1着"],
)
def test_apply_color_rules_grade_finish_within_matches(value: str) -> None:
    """grade_finish_withinは着順がグレード別上限以内のときPatternFillを返す。"""
    rules = [_rule("grade_finish_within", {"G1": 9, "G2": 2, "G3": 1}, "yellow")]
    result = apply_color_rules(value, rules)
    assert isinstance(result, PatternFill)


@pytest.mark.parametrize(
    "value",
    ["G1 10着", "G2 3着", "G3 2着", "非重賞 1着", "不正な値"],
)
def test_apply_color_rules_grade_finish_within_does_not_match(value: str) -> None:
    """grade_finish_withinは着順上限超過・対象外グレード・不正形式のときNoneを返す。"""
    rules = [_rule("grade_finish_within", {"G1": 9, "G2": 2, "G3": 1}, "yellow")]
    assert apply_color_rules(value, rules) is None
